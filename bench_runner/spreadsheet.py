from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from bench_runner.error_analysis import classify_spreadsheet_error, insert_error_type
from bench_runner.prompts import SpreadsheetPromptContext, SYSTEM_PROMPTS, build_spreadsheet_user_prompt

_ROOT = Path(__file__).resolve().parents[1]


def _repo_path(path: str | Path) -> Path:
    path = Path(path)
    return path if path.is_absolute() else (_ROOT / path)


def _safe_name(value: str) -> str:
    safe = re.sub(r"[^\w.\-]+", "_", str(value), flags=re.ASCII)
    return safe.strip("_") or "model"


def _dataset_result_name(test_data_path: str | Path) -> str:
    return f"results_{Path(test_data_path).stem}.json"


def _relative_to(path: Path, base: Path) -> str:
    try:
        return str(path.relative_to(base)).replace("\\", "/")
    except ValueError:
        return str(path)


def _dumps_results(payload: Dict[str, Any]) -> str:
    if "questions" not in payload or not isinstance(payload["questions"], list):
        return json.dumps(payload, ensure_ascii=False, indent=2)

    head = {key: value for key, value in payload.items() if key != "questions"}
    lines = ["{"]
    for key, value in head.items():
        lines.append(f'  {json.dumps(key, ensure_ascii=False)}: {json.dumps(value, ensure_ascii=False)},')
    lines.append('  "questions": [')
    for idx, question in enumerate(payload["questions"]):
        suffix = "," if idx < len(payload["questions"]) - 1 else ""
        lines.append(f"    {json.dumps(question, ensure_ascii=False)}{suffix}")
    lines.append("  ]")
    lines.append("}")
    return "\n".join(lines)


def _read_json(path: str | Path) -> Any:
    with _repo_path(path).open("r", encoding="utf-8") as f:
        return json.load(f)


def _write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _resolve_workbook_path(raw_path: str) -> Path:
    return _repo_path(raw_path)


def _number_workbook_path(formula_workbook: Path) -> Path:
    name = formula_workbook.name.replace("_solution_formula.xlsx", "_solution_number.xlsx")
    return formula_workbook.with_name(name)


def _formula_workbook_path(template_workbook: Path) -> Path:
    name = template_workbook.name.replace("_workbook.xlsx", "_solution_formula.xlsx")
    return template_workbook.with_name(name)


def _question_id(question: Dict[str, Any]) -> int:
    return int(question["id"])


def _dependency_ids(question: Dict[str, Any]) -> List[int]:
    depends_on = question.get("depends_on") or []
    if isinstance(depends_on, (str, int)):
        depends_on = [depends_on]
    ids: List[int] = []
    for item in depends_on:
        try:
            ids.append(int(item))
        except (TypeError, ValueError):
            continue
    return ids


def _dependency_formula_input(output_root: Path, question: Dict[str, Any]) -> Optional[Tuple[int, Path]]:
    for dependency_id in reversed(_dependency_ids(question)):
        dependency_formula = output_root / f"ID{dependency_id:03d}" / f"ID{dependency_id:03d}_formula.xlsx"
        if dependency_formula.exists():
            return dependency_id, dependency_formula
    return None


def _answer_positions(answer_position: Any) -> List[str]:
    if isinstance(answer_position, list):
        return [str(item) for item in answer_position]
    if answer_position is None:
        return []
    return [str(answer_position)]


def _split_sheet_range(address: str) -> Tuple[str, str]:
    if "!" not in address:
        raise ValueError(f"Answer position must include a sheet name: {address}")
    sheet, cell_range = address.rsplit("!", 1)
    sheet = sheet.strip()
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1].replace("''", "'")
    return sheet, cell_range.strip()


def _quote_sheet(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def _expand_position(address: str) -> List[Tuple[str, str]]:
    sheet, cell_range = _split_sheet_range(address)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cells = []
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cells.append((sheet, f"{get_column_letter(col)}{row}"))
    return cells


def _target_cells(answer_position: Any) -> List[Tuple[str, str]]:
    cells: List[Tuple[str, str]] = []
    seen = set()
    for position in _answer_positions(answer_position):
        for cell in _expand_position(position):
            key = (cell[0], cell[1].upper())
            if key not in seen:
                seen.add(key)
                cells.append((cell[0], cell[1].upper()))
    return cells


def _cell_key(sheet: str, cell: str) -> str:
    return f"{_quote_sheet(sheet)}!{cell.upper()}"


def _workbook_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as workbook_file:
        for chunk in iter(lambda: workbook_file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _display_value(value: Any, max_len: int = 120) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    if len(text) > max_len:
        return text[: max_len - 3] + "..."
    return text


def _cell_snapshot(cell: Any) -> str:
    address = cell.coordinate
    value = cell.value
    parts = [address]
    if isinstance(value, str) and value.startswith("="):
        parts.append(f"formula={_display_value(value, 220)}")
    elif value is not None:
        parts.append(f"value={_display_value(value)}")
    if cell.number_format and cell.number_format != "General":
        parts.append(f"format={_display_value(cell.number_format, 60)}")
    if cell.comment and cell.comment.text:
        parts.append(f"comment={_display_value(cell.comment.text, 120)}")
    return " | ".join(parts)


def _sheet_nonempty_rows(sheet: Any) -> Tuple[int, List[str]]:
    rows: List[str] = []
    count = 0
    for row in sheet.iter_rows():
        cells = []
        for cell in row:
            if cell.value is None and not cell.comment:
                continue
            count += 1
            cells.append(_cell_snapshot(cell))
        if cells:
            rows.append(f"R{row[0].row}: " + " ; ".join(cells))
    return count, rows


def _range_context_lines(workbook: Any, address: str, radius: int = 2) -> List[str]:
    sheet_name, cell_range = _split_sheet_range(address)
    if sheet_name not in workbook.sheetnames:
        return [f"{address}: sheet not found"]
    sheet = workbook[sheet_name]
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    min_col = max(1, min_col - radius)
    min_row = max(1, min_row - radius)
    max_col = min(sheet.max_column, max_col + radius)
    max_row = min(sheet.max_row, max_row + radius)
    lines = [f"{_quote_sheet(sheet_name)}!{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"]
    for row_idx in range(min_row, max_row + 1):
        cells = []
        for col_idx in range(min_col, max_col + 1):
            cell = sheet.cell(row_idx, col_idx)
            if cell.value is not None:
                cells.append(_cell_snapshot(cell))
        if cells:
            lines.append(f"R{row_idx}: " + " ; ".join(cells))
    return lines


def _workbook_snapshot_text(workbook_path: Path, answer_position: Any) -> str:
    max_chars = int(os.environ.get("SPREADSHEET_WORKBOOK_SNAPSHOT_MAX_CHARS", "450000"))
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        lines = [
            "WORKBOOK_SNAPSHOT_BEGIN",
            f"file_name: {workbook_path.name}",
            f"file_sha256: {_workbook_sha256(workbook_path)}",
            f"sheet_count: {len(workbook.sheetnames)}",
            "sheets: " + ", ".join(workbook.sheetnames),
        ]

        defined_names = []
        for name in workbook.defined_names:
            try:
                defined_names.append(f"{name.name}={name.attr_text}")
            except AttributeError:
                defined_names.append(str(name))
        if defined_names:
            lines.append("named_ranges: " + " ; ".join(defined_names))

        lines.append("FULL_WORKBOOK_CELLS_BEGIN")
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            nonempty_count, row_lines = _sheet_nonempty_rows(sheet)
            lines.append(
                f"SHEET {sheet_name} | used_range=A1:{get_column_letter(sheet.max_column)}{sheet.max_row} "
                f"| non_empty_cells={nonempty_count}"
            )
            merged_ranges = [str(item) for item in sheet.merged_cells.ranges]
            if merged_ranges:
                lines.append("merged_ranges: " + ", ".join(merged_ranges))
            if getattr(sheet, "tables", None):
                table_names = list(sheet.tables.keys())
                if table_names:
                    lines.append("tables: " + ", ".join(table_names))
            lines.extend(row_lines)
        lines.append("FULL_WORKBOOK_CELLS_END")
        lines.append("WORKBOOK_SNAPSHOT_END")
        snapshot = "\n".join(lines)
        if len(snapshot) <= max_chars:
            return snapshot
        return (
            snapshot[:max_chars]
            + "\nWORKBOOK_SNAPSHOT_TRUNCATED: increase SPREADSHEET_WORKBOOK_SNAPSHOT_MAX_CHARS to include more workbook content."
        )
    finally:
        workbook.close()


def _strip_code_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        lines = text.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip().startswith("```"):
            lines = lines[:-1]
        return "\n".join(lines).strip()
    return text


def _extract_python_code(response: str) -> str:
    if not response:
        return ""
    match = re.search(r"```(?:python|py)?\s*(.*?)```", response, flags=re.IGNORECASE | re.DOTALL)
    if match:
        return match.group(1).strip()
    return _strip_code_fences(response).strip()


def _run_generated_spreadsheet_code(
    *,
    code: str,
    question_dir: Path,
    template_path: Path,
    output_workbook: Path,
    timeout_seconds: int = 240,
) -> Dict[str, Any]:
    question_dir = question_dir.resolve()
    template_path = template_path.resolve()
    output_workbook = output_workbook.resolve()
    question_dir.mkdir(parents=True, exist_ok=True)

    input_path = question_dir / "input.xlsx"
    expected_output = question_dir / "output.xlsx"
    script_path = question_dir / "generated_spreadsheet.py"

    shutil.copy2(template_path, input_path)
    _write_text(script_path, code)

    env = os.environ.copy()
    env["INPUT_WORKBOOK"] = str(input_path.resolve())
    env["OUTPUT_WORKBOOK"] = str(expected_output.resolve())
    result = subprocess.run(
        [sys.executable, str(script_path.name)],
        cwd=str(question_dir),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        timeout=timeout_seconds,
        check=False,
        env=env,
    )

    created_path = expected_output if expected_output.exists() else output_workbook
    if created_path.exists():
        shutil.copy2(created_path, output_workbook)

    return {
        "return_code": result.returncode,
        "stdout": result.stdout,
        "stderr": result.stderr,
        "script": str(script_path.resolve()),
        "input_workbook": str(input_path.resolve()),
        "expected_output": str(expected_output.resolve()),
        "output_created": output_workbook.exists(),
    }


def _parse_cell_ref(raw_ref: str) -> Tuple[Optional[str], str]:
    raw_ref = raw_ref.strip().strip("`")
    if "!" in raw_ref:
        sheet, cell = _split_sheet_range(raw_ref)
        return sheet, cell.upper()
    return None, raw_ref.upper()


def _parse_cell_value(raw_value: str) -> Any:
    value = raw_value.strip()
    if not value:
        return None
    if value.startswith("="):
        return value
    if (value.startswith('"') and value.endswith('"')) or (value.startswith("'") and value.endswith("'")):
        return value[1:-1]
    try:
        if re.fullmatch(r"[-+]?\d+", value):
            return int(value)
        if re.fullmatch(r"[-+]?(?:\d+\.\d*|\.\d+)(?:[eE][-+]?\d+)?", value) or re.fullmatch(r"[-+]?\d+[eE][-+]?\d+", value):
            return float(value)
    except ValueError:
        pass
    return value


def _parse_llm_cells(response: str, targets: List[Tuple[str, str]]) -> Dict[Tuple[str, str], Any]:
    target_by_cell: Dict[str, List[str]] = {}
    for sheet, cell in targets:
        target_by_cell.setdefault(cell.upper(), []).append(sheet)

    parsed: Dict[Tuple[str, str], Any] = {}
    text = _strip_code_fences(response)
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line or ":" not in line:
            continue
        raw_ref, raw_value = line.split(":", 1)
        sheet, cell = _parse_cell_ref(raw_ref)
        if not re.fullmatch(r"\$?[A-Z]{1,3}\$?\d+", cell):
            continue
        cell = cell.replace("$", "").upper()
        if sheet is None:
            sheets = target_by_cell.get(cell, [])
            if len(sheets) != 1:
                continue
            sheet = sheets[0]
        key = (sheet, cell)
        if key in targets:
            parsed[key] = _parse_cell_value(raw_value)
    return parsed


def _load_cell_map(path: Path, targets: List[Tuple[str, str]], data_only: bool) -> Dict[Tuple[str, str], Any]:
    workbook = load_workbook(path, data_only=data_only)
    values: Dict[Tuple[str, str], Any] = {}
    try:
        for sheet, cell in targets:
            if sheet in workbook.sheetnames:
                values[(sheet, cell)] = workbook[sheet][cell].value
            else:
                values[(sheet, cell)] = None
    finally:
        workbook.close()
    return values


def _value_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _formula_match(candidate: Any, reference: Any) -> bool:
    return _value_text(candidate) == _value_text(reference)


def _numeric_match(candidate: Any, reference: Any, tolerance: float = 0.01) -> bool:
    if candidate is None and reference is None:
        return True
    try:
        cand = float(candidate)
        ref = float(reference)
        if ref == 0:
            return abs(cand) <= 1e-9
        return abs(cand - ref) / abs(ref) <= tolerance
    except (TypeError, ValueError):
        return _value_text(candidate) == _value_text(reference)


def _find_soffice() -> Optional[str]:
    env_path = os.getenv("LIBREOFFICE_PATH") or os.getenv("SOFFICE_PATH")
    if env_path and Path(env_path).exists():
        return env_path
    found = shutil.which("soffice") or shutil.which("libreoffice")
    if found:
        return found
    for candidate in (
        r"C:\Program Files\LibreOffice\program\soffice.exe",
        r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    ):
        if Path(candidate).exists():
            return candidate
    return None


def _recalculate_with_com(path: Path) -> bool:
    helper_code = r'''
import sys
import pythoncom
import win32com.client as win32

path = sys.argv[1]
prog_ids = ["ket.Application", "et.Application", "Ex" + "cel.Application"]
pythoncom.CoInitialize()
for prog_id in prog_ids:
    app = None
    workbook = None
    try:
        app = win32.DispatchEx(prog_id)
        app.Visible = False
        try:
            app.DisplayAlerts = False
        except Exception:
            pass
        workbook = app.Workbooks.Open(path)
        for method_name in ("CalculateFullRebuild", "CalculateFull", "Calculate"):
            try:
                getattr(app, method_name)()
                break
            except Exception:
                continue
        workbook.Save()
        workbook.Close(SaveChanges=True)
        app.Quit()
        print(prog_id)
        sys.exit(0)
    except Exception:
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if app is not None:
                app.Quit()
        except Exception:
            pass
print("No COM spreadsheet engine available")
sys.exit(1)
'''
    with tempfile.TemporaryDirectory() as temp_dir:
        helper_path = Path(temp_dir) / "recalculate_com.py"
        helper_path.write_text(helper_code, encoding="utf-8")
        result = subprocess.run(
            [sys.executable, str(helper_path), str(path.resolve())],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=180,
            check=False,
        )
        return result.returncode == 0


def _recalculate_workbook(path: Path) -> bool:
    try:
        if _recalculate_with_com(path):
            return True
    except Exception:
        pass

    soffice = _find_soffice()
    if not soffice:
        return False
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir) / path.name
        shutil.copy2(path, temp_path)
        out_dir = Path(temp_dir) / "out"
        out_dir.mkdir()
        result = subprocess.run(
            [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(out_dir), str(temp_path)],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=120,
            check=False,
        )
        converted = out_dir / path.name
        if result.returncode == 0 and converted.exists():
            shutil.copy2(converted, path)
            return True
    return False


def _apply_response_to_workbook(template_path: Path, output_path: Path, values: Dict[Tuple[str, str], Any]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template_path, output_path)
    workbook = load_workbook(output_path)
    try:
        for (sheet, cell), value in values.items():
            if sheet not in workbook.sheetnames:
                continue
            workbook[sheet][cell] = value
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"
        workbook.save(output_path)
    finally:
        workbook.close()


def _create_number_workbook(formula_workbook: Path, number_workbook: Path) -> bool:
    number_workbook.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(formula_workbook, number_workbook)
    return _recalculate_workbook(number_workbook)


def _score_workbook(
    candidate_number_path: Path,
    number_path: Path,
    targets: List[Tuple[str, str]],
    calculation_available: bool,
) -> Dict[str, Any]:
    candidate_number = _load_cell_map(candidate_number_path, targets, data_only=True) if candidate_number_path.exists() else {}
    reference_number = _load_cell_map(number_path, targets, data_only=True)

    cell_scores = []
    number_hits = 0
    for sheet, cell in targets:
        key = (sheet, cell)
        cell_score = 1.0 if _numeric_match(candidate_number.get(key), reference_number.get(key)) else 0.0
        number_score_source = "calculated" if calculation_available else "cached_value"
        number_hits += int(cell_score)
        cell_scores.append(
            {
                "cell": _cell_key(sheet, cell),
                "score": cell_score,
                "number_score_source": number_score_source,
                "candidate_number": candidate_number.get(key),
                "reference_number": reference_number.get(key),
            }
        )

    total = len(targets)
    score = number_hits / total if total else 0.0
    return {
        "status": "scored",
        "calculation_available": calculation_available,
        "numerical_fallback_used": not calculation_available,
        "cell_count": total,
        "number_correct_cells": number_hits,
        "score": score,
        "cell_scores": cell_scores,
    }


def _iter_questions(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, dict):
        return list(data.get("questions", []))
    return [item for item in data if isinstance(item, dict) and item.get("template")]


def _select_questions(
    questions: List[Dict[str, Any]],
    limit: Optional[int],
    group_ids: List[int],
    question_ids: Optional[List[int]] = None,
) -> List[Dict[str, Any]]:
    selected = questions
    if question_ids:
        wanted_questions = set(question_ids)
        selected = [q for q in selected if int(q.get("id", q.get("question_num", 0)) or 0) in wanted_questions]
    if group_ids:
        wanted = set(group_ids)
        selected = [q for q in selected if int(q.get("group_id", 0)) in wanted]
    if limit and limit > 0:
        selected = selected[:limit]
    return selected


def run_spreadsheet_benchmark(
    *,
    model: Any,
    model_key: str,
    test_data_path: str,
    output_dir: str,
    limit: Optional[int],
    group_ids: List[int],
    question_ids: Optional[List[int]] = None,
) -> Dict[str, Any]:
    data = _read_json(test_data_path)
    questions = _select_questions(_iter_questions(data), limit, group_ids, question_ids)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_root = _repo_path(output_dir) / f"{_safe_name(model_key)}_{timestamp}"
    artifact_root = run_root / "LV3.1"
    artifact_root.mkdir(parents=True, exist_ok=True)
    result_path = run_root / _dataset_result_name(test_data_path)

    rows = []

    def write_current_summary() -> Dict[str, Any]:
        summary = {
            "model": model_key,
            "dataset": Path(test_data_path).stem,
            "timestamp": timestamp,
            "total_questions": len(rows),
            "average_score": round(sum(float(row["score"] or 0.0) for row in rows) / len(rows), 6) if rows else 0.0,
            "questions": rows,
        }
        _write_text(result_path, _dumps_results(summary))
        return summary

    for question in questions:
        question_num = _question_id(question)
        item_id = f"ID{question_num:03d}"
        item_dir = artifact_root / item_id
        item_dir.mkdir(parents=True, exist_ok=True)

        template_path = _resolve_workbook_path(question["template"])
        dependency_input = _dependency_formula_input(artifact_root, question)
        dependency_input_id = dependency_input[0] if dependency_input else None
        input_workbook_path = dependency_input[1] if dependency_input else template_path
        formula_path = _formula_workbook_path(template_path)
        number_path = _number_workbook_path(formula_path)
        answer_position = question["answer_position"]
        targets = _target_cells(answer_position)
        formula_workbook = item_dir / f"{item_id}_formula.xlsx"
        number_workbook = item_dir / f"{item_id}_number.xlsx"
        workbook_snapshot = _workbook_snapshot_text(input_workbook_path, answer_position)
        snapshot_file = item_dir / f"{item_id}_workbook_snapshot.txt"
        _write_text(snapshot_file, workbook_snapshot)

        ctx = SpreadsheetPromptContext(
            task_id=item_id,
            question_id=question_num,
            group_id=str(question.get("group_id", "")),
            instruction=str(question.get("question", "")),
            background="",
            template_workbook=str(template_path),
            answer_position=answer_position,
            workbook_snapshot=workbook_snapshot,
            knowledge_tags=[],
            dependency_questions=(
                [f"input.xlsx is the completed formula workbook from dependency ID{dependency_input_id:03d}."]
                if dependency_input_id is not None
                else
                ["No dependency result workbook was found; input.xlsx is the original template workbook."]
                if _dependency_ids(question)
                else []
            ),
        )
        user_prompt = build_spreadsheet_user_prompt(ctx)

        raw_response = ""
        score: Dict[str, Any]
        error = None
        generated_code = ""
        run_result: Optional[Dict[str, Any]] = None
        response_time_seconds = None
        execution_time_seconds = None
        input_token = None
        try:
            response_started = time.perf_counter()
            response = model.generate_with_retry_details(
                user_prompt,
                retry_times=3,
                system_prompt=SYSTEM_PROMPTS["spreadsheet"],
            )
            response_time_seconds = round(time.perf_counter() - response_started, 2)
            input_token = response.get("input_tokens")
            raw_response = response.get("raw_response", "") or ""
            final_response = response.get("final_response", "") or raw_response
            generated_code = _extract_python_code(final_response)
            with tempfile.TemporaryDirectory() as temp_dir:
                run_result = _run_generated_spreadsheet_code(
                    code=generated_code,
                    question_dir=Path(temp_dir),
                    template_path=input_workbook_path,
                    output_workbook=formula_workbook,
                )
                if run_result["return_code"] != 0:
                    stderr_tail = (run_result.get("stderr") or "").strip()[-2000:]
                    raise RuntimeError(
                        f"Generated spreadsheet code failed with exit code {run_result['return_code']}. "
                        f"stderr: {stderr_tail}"
                    )
                if not run_result["output_created"]:
                    raise FileNotFoundError("Generated spreadsheet code did not create output.xlsx")
            recalc_started = time.perf_counter()
            calculation_available = _create_number_workbook(formula_workbook, number_workbook)
            execution_time_seconds = round(time.perf_counter() - recalc_started, 2)
            score = _score_workbook(
                number_workbook,
                number_path,
                targets,
                calculation_available=calculation_available,
            )
        except Exception as exc:
            error = str(exc)
            score = {
                "status": "error",
                "calculation_available": False,
                "cell_count": len(targets),
                "score": 0.0,
                "cell_scores": [],
            }

        result_entry = {
            "id": question_num,
            "source": question.get("source", ""),
            "raw_response": raw_response,
            "input_file": input_workbook_path.name,
            "input_source": f"dependency_formula:ID{dependency_input_id:03d}" if dependency_input_id is not None else "template",
            "workbook_snapshot_file": _relative_to(snapshot_file, run_root),
            "score": score.get("score", 0.0),
            "formula_file": _relative_to(formula_workbook, run_root),
            "number_file": _relative_to(number_workbook, run_root),
            "input_token": input_token,
            "response_time_seconds": response_time_seconds,
            "execution_time_seconds": execution_time_seconds,
            "error": error,
        }
        result_entry = insert_error_type(result_entry, classify_spreadsheet_error(result_entry, run_root=run_root))
        rows.append(result_entry)
        summary = write_current_summary()

    summary = write_current_summary()
    print(result_path)
    return summary
